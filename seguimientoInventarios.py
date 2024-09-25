from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Crear un nuevo libro de trabajo
wb = Workbook()

# Crear una hoja de cálculo para el inventario
sheet = wb.active
sheet.title = 'Inventario'

# Encabezados
sheet['A1'] = 'Producto'
sheet['B1'] = 'Cantidad Disponible'
sheet['C1'] = 'Ubicación'

# Aplicar formato a los encabezados
header_font = Font(bold=True)
for col in range(1, 4):
    cell = sheet.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Ejemplo de datos (puedes modificar esta parte según tus necesidades)
inventario = [
    ('Producto A', 500, 'Estante 1'),
    ('Producto B', 700, 'Estante 2'),
    ('Producto C', 300, 'Estante 3'),
    ('Producto D', 400, 'Estante 1'),
]

# Escribir datos en el archivo
for row_idx, row_data in enumerate(inventario, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx).value = cell_data

# Ajustar ancho de columnas automáticamente
for col in sheet.iter_cols():
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Guardar el libro de trabajo
wb.save('inventario.xlsx')

# Cerrar el libro de trabajo
wb.close()

print("Archivo 'inventario.xlsx' creado correctamente.")
