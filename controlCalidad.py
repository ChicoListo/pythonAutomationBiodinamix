from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, Reference

# Crear un nuevo libro de trabajo
wb = Workbook()

# Crear una hoja de cálculo para el control de calidad
sheet = wb.active
sheet.title = 'Control de Calidad'

# Encabezados
sheet['A1'] = 'Fecha'
sheet['B1'] = 'Producto'
sheet['C1'] = 'Cantidad Producida'
sheet['D1'] = 'Defectos'
sheet['E1'] = 'Porcentaje Defectuoso'

# Aplicar formato a los encabezados
header_font = Font(bold=True)
for col in range(1, 6):
    cell = sheet.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Ejemplo de datos (puedes modificar esta parte según tus necesidades)
data = [
    ('2024-07-09', 'Producto A', 1000, 20),
    ('2024-07-09', 'Producto B', 1500, 30),
    ('2024-07-09', 'Producto C', 800, 15),
    ('2024-07-09', 'Producto D', 1200, 25),
]

# Escribir datos en el archivo
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx).value = cell_data

# Calcular porcentaje defectuoso y aplicar formato
for row in range(2, len(data) + 2):
    cantidad_producida = sheet.cell(row=row, column=3).value
    defectos = sheet.cell(row=row, column=4).value
    if cantidad_producida and defectos:
        porcentaje_defectuoso = defectos / cantidad_producida * 100
        sheet.cell(row=row, column=5).value = porcentaje_defectuoso
        sheet.cell(row=row, column=5).number_format = '0.00%'

# Agregar gráfico de pastel
chart = PieChart()
labels = Reference(sheet, min_col=2, min_row=2, max_row=len(data) + 1)
data = Reference(sheet, min_col=5, min_row=1, max_row=len(data) + 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = 'Porcentaje Defectuoso por Producto'

# Ubicar el gráfico dentro de la hoja de cálculo
sheet.add_chart(chart, "G1")

# Ajustar ancho de columnas automáticamente
for col in sheet.iter_cols():
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Guardar el libro de trabajo
wb.save('control_calidad_con_grafico.xlsx')

# Cerrar el libro de trabajo
wb.close()

print("Archivo 'control_calidad_con_grafico.xlsx' creado correctamente.")

