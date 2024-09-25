#Este script analisa el archivo de inspecciones de calidad mostrando la cantidad de productos inspeccionados, productos aceptables, productos defectuosos y porcentaje de productos defectuosos.
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Cargar el libro de trabajo existente
wb = load_workbook('inspecciones_calidad.xlsx')

# Seleccionar la hoja de cálculo de inspecciones de calidad
sheet = wb['Inspecciones de Calidad']

# Calcular el número total de productos inspeccionados
total_productos = sheet.max_row - 1  # Restar 1 para excluir la fila de encabezados

# Calcular la cantidad de productos aceptables y defectuosos
productos_aceptables = 0
productos_defectuosos = 0

for row in range(2, sheet.max_row + 1):
    resultado = sheet.cell(row=row, column=column_index_from_string('D')).value
    if resultado == 'Aceptable':
        productos_aceptables += 1
    elif resultado == 'Defectuoso':
        productos_defectuosos += 1

# Calcular el porcentaje de productos defectuosos
porcentaje_defectuosos = (productos_defectuosos / total_productos) * 100

# Mostrar resultados
print(f"Total de productos inspeccionados: {total_productos}")
print(f"Productos aceptables: {productos_aceptables}")
print(f"Productos defectuosos: {productos_defectuosos}")
print(f"Porcentaje de productos defectuosos: {porcentaje_defectuosos:.2f}%")

# Cerrar el libro de trabajo
wb.close()
