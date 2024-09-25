from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# Crear un nuevo libro de trabajo
wb = Workbook()

# Crear una hoja de cálculo para las inspecciones de calidad
sheet = wb.active
sheet.title = 'Inspecciones de Calidad'

# Encabezados
sheet['A1'] = 'Fecha'
sheet['B1'] = 'Producto'
sheet['C1'] = 'Cantidad'
sheet['D1'] = 'Resultado'
sheet['E1'] = 'Observaciones'

# Aplicar formato a los encabezados
header_font = Font(bold=True)
for col in range(1, 6):
    cell = sheet.cell(row=1, column=col)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Ejemplo de datos (puedes modificar esta parte según tus necesidades)
data = [
    ('2024-07-09', 'Producto A', 100, 'Aceptable', 'Ninguna'),
    ('2024-07-09', 'Producto B', 150, 'Defectuoso', 'Falta de etiqueta'),
    ('2024-07-09', 'Producto C', 80, 'Aceptable', 'Ninguna'),
    ('2024-07-09', 'Producto D', 120, 'Aceptable', 'Ninguna'),
]

# Escribir datos en el archivo
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, cell_data in enumerate(row_data, start=1):
        sheet.cell(row=row_idx, column=col_idx).value = cell_data

# Ajustar ancho de columnas automáticamente
for col in sheet.iter_cols():
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column].width = adjusted_width

# Guardar el libro de trabajo
wb.save('inspecciones_calidad.xlsx')

# Cerrar el libro de trabajo
wb.close()

print("Archivo 'inspecciones_calidad.xlsx' creado correctamente.")
